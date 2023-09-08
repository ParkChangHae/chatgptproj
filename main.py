import re
import time
from datetime import datetime

import numpy
import openai
import openpyxl
import random

from openpyxl import load_workbook
from openpyxl.styles import Alignment

openai.api_key = "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"

def load_excel_data(filename, sheet_name):
    # 엑셀 파일을 불러옴
    wb = openpyxl.load_workbook(filename, data_only=True)
    # 첫 번째 시트를 선택
    sheet = wb[sheet_name]
    #sheet   = wb.active

    # 데이터를 읽어서 딕셔너리 리스트로 변환
    data = []
    row_data = {}

    column_name = []

    #첫 항목명들을 읽어와서 우선 저장
    for row in sheet.iter_rows(min_row=2, min_col=2, max_row=2):
        for row_element in row:
            if(row_element.value != None):
                column_name.append(row_element.value)

    for row in sheet.iter_rows(min_row=3, min_col=2):  #B3' 셀부터 시작
        row_data = {}

        for i in range(0, len(column_name)):
            row_data[column_name[i]] = row[i].value

        if(row[0].value != None):
            data.append(row_data)

    return data


def save_excel_data(problem_info_list, temperature, top_p):

    wb = load_workbook('./output/output.xlsx')

    # 워크시트 선택
    ws = wb.active

    # 데이터를 리스트 형태로 준비합니다.
    # 예시: 각 행의 데이터를 리스트로 표현

    for problem in problem_info_list:

        current_row = int(problem['No.'].replace('.','')) - 16

        ws.cell(row = current_row, column = 2, value = problem['problem']['질문']).alignment = Alignment(wrapText=True, vertical='center')

        ws.cell(row = current_row, column = 3, value = problem['problem']['문제 본문']).alignment = Alignment(wrapText=True, vertical='center')

        if('해설' in problem['problem'].keys()):
            ws.cell(row = current_row, column = 7, value = problem['problem']['해설']).alignment = Alignment(wrapText=True, vertical='center')

        if('주요 어휘' in problem['problem'].keys()):
            ws.cell(row = current_row, column = 8, value = problem['problem']['주요 어휘']).alignment = Alignment(wrapText=True, vertical='center')

        if( problem['theme'] == "틀린 어법 찾기" or problem['theme'] == "어휘" or problem['theme'] == "전체 흐름과 관계 없는 문장 파악"  ):
            pass

        elif (problem['theme'] == "글의 순서 추론"):
            ws.cell(row = current_row, column = 5, value = problem['problem']['주어진 글']).alignment = Alignment(wrapText=True, vertical='center')
            ws.cell(row = current_row, column = 4, value = problem['problem']['선택지']).alignment = Alignment( wrapText=True, vertical='center')

        elif (problem['theme'] == "문장 삽입"):
            ws.cell(row = current_row, column = 5, value = problem['problem']['주어진 문장']).alignment = Alignment(wrapText=True, vertical='center')

        elif (problem['theme'] == "문장 요약"):
            ws.cell(row = current_row, column = 4, value = problem['problem']['선택지']).alignment = Alignment(wrapText=True, vertical='center')
            ws.cell(row = current_row, column = 5, value = problem['problem']['요약 문장']).alignment = Alignment(wrapText=True, vertical='center')

        else :
            ws.cell(row = current_row, column = 4, value = problem['problem']['선택지']).alignment = Alignment(wrapText=True, vertical='center')

        try :

            strings = ['\n'.join([f'{k}: {v}' for k, v in d.items()]) for d in problem['few_shot_message']]
            result_string = '\n'.join(strings)

            ws.cell(row=current_row, column=11, value=result_string)

        except:
            pass

    # 현재 날짜와 시간을 가져옵니다.
    now = datetime.now()

    # 날짜와 시간을 문자열로 변환합니다. 여기서는 '년-월-일 시:분' 형식을 사용했습니다.
    now_str = now.strftime('%Y%m%d_%H%M')

    # 엑셀 파일 저장
    wb.save(f'./output/output_temp={temperature}_top_p={top_p}_{now_str}.xlsx')


def make_problem_example_set(data):

    random_items = random.sample(data, 5)

    message_list = [
        {"role": "system", "content": "You are an assistant who kindly solves or produces English problems."},
        ]

    for data in random_items:

        if( data['해설'] != None and data['어휘'] != None ):

            if( data['단원명'] == "틀린 어법 찾기" or data['단원명'] == "어휘"
                    or data['단원명'] == "전체 흐름과 관계 없는 문장 파악"  ):

                main_string = "다음은 {}을 주제로 한 영어 문항, 문제 본문 해석과 문항의 해설이야.\n질문 : {}\n문제 본문 : {}\n본문 단어 설명 : {}\n해설 : {}\n주요 어휘 : {}".format(
                    data['단원명'],data['발문'],data['문제 본문'],data['단어 설명'],data['해설'],data['어휘'])

            elif( data['단원명'] == "글의 순서 추론" ):
                main_string = "다음은 {}을 주제로 한 영어 문항과 문항의 해설이야. (단, 본문은 (A), (B), (C) 최대 3개로만 구성)\n질문 : {}\n주어진 글 : {}\n문제 본문 : {}\n본문 단어 설명 : {}\n선택지 : {}\n해설 : {}\n주요 어휘 : {}".format(
                    data['단원명'], data['발문'],data['주어진 글'], data['문제 본문'],data['단어 설명'], data['선지'], data['해설'], data['어휘'])

            elif( data['단원명'] == "문장 삽입" ):
                main_string = "다음은 {}을 주제로 한 영어 문항과 문항의 해설이야.\n질문 : {}\n주어진 문장 : {}\n문제 본문 : {}\n본문 단어 설명 : {}\n해설 : {}\n주요 어휘 : {}".format(
                    data['단원명'], data['발문'],data['주어진 문장'], data['문제 본문'],data['단어 설명'], data['해설'], data['어휘'])

            elif( data['단원명'] == "문장 요약" ):
                main_string = "다음은 {}을 주제로 한 영어 문항과 문항의 해설이야.\n질문 : {}\n문제 본문 : {}\n본문 단어 설명 : {}\n요약 문장 : {}\n선택지 : {}\n해설 : {}\n주요 어휘 : {}".format(
                    data['단원명'], data['발문'], data['문제 본문'],data['단어 설명'], data['요약 문장'],data['선지'], data['해설'], data['어휘'])

            else:
                main_string = "다음은 {}을 주제로 한 영어 문항, 문제 본문 해석과 문항의 해설이야.\n질문 : {}\n문제 본문 : {}\n본문 단어 설명 : {}\n선택지 : {}\n해설 : {}\n주요 어휘 : {}".format(
                    data['단원명'],data['발문'],data['문제 본문'],data['단어 설명'],data['선지'],data['해설'],data['어휘'])

            message_list.append({"role" : "user", "content" : main_string})

    return message_list

def make_single_problem(problem_theme, temperature, top_p):

    message_template = {"role": "user", "content": f"위에 보여준 예시 영어 {problem_theme} 문항들의 형식을 참고해서, "
                                                   f"{problem_theme} 유형의 영어 문항과 해설을 하나 만들어줘. 내가 예시로 든 문항들의 글감을 따라하지는 말고, 최대한 다양한 분야에서 다양한 글감으로 지문을 구성해줘. '질문 : '으로 시작해줘"}

    data = load_excel_data('./raw_data.xlsx', problem_theme)

    is_not_FATAL_ERROR = True
    is_not_success = True

    while is_not_FATAL_ERROR:

        few_shot_message_list = make_problem_example_set(data)
        few_shot_message_list.insert(len(few_shot_message_list) - 1, message_template)
        few_shot_message_list[-1]['role'] = 'assistant'
        few_shot_message_list[-1]['content'] = '질문 : ' + few_shot_message_list[-1]['content'].split('질문 : ')[-1]
        few_shot_message_list.append(message_template)

        while is_not_success:

            try :

                '''
                                response = openai.ChatCompletion.create(

                    model="gpt-3.5-turbo-16k-0613",
                    messages=few_shot_message_list,
                    temperature=temperature,
                    max_tokens=4096,
                    top_p=top_p,
                    frequency_penalty=0.0,
                    presence_penalty=0.0,
                )
                '''

                response = openai.ChatCompletion.create(

                    model="gpt-4-0613",
                    messages=few_shot_message_list,
                    temperature=temperature,
                    max_tokens=4096,
                    top_p=top_p,
                    frequency_penalty=0.0,
                    presence_penalty=0.0,
                )

                is_not_success = False
                is_not_FATAL_ERROR = False

            except Exception as e:

                print(e)
                is_not_success = True

                if(len(few_shot_message_list) >= 4):
                    few_shot_message_list.pop(1)
                else:
                    is_not_FATAL_ERROR = True





    print("problem_theme : " + problem_theme + " message_cnt : " + str(len(few_shot_message_list)))

    return response['choices'][0]['message']['content'], few_shot_message_list

def make_testpaper(temperature, top_p):

    problem_No_theme_dict_list = [
    {"No.": "18.", "theme": "글의 목적 파악"},
    {"No.": "19.", "theme": "심경 분위기 파악"},
    {"No.": "20.", "theme": "필자의 주장 파악"},
    {"No.": "21.", "theme": "괄호 친 문장의 의미 파악"},
    {"No.": "22.", "theme": "글의 요지 파악"},
    {"No.": "23.", "theme": "글의 주제 파악"},
    {"No.": "24.", "theme": "글의 제목 추론"},
    {"No.": "25.", "theme": "도표 내용 일치 불일치"},
    {"No.": "26.", "theme": "대상 불일치파악"},
    {"No.": "27.", "theme": "안내문 불일치 파악"},
    {"No.": "28.", "theme": "안내문 일치 파악"},
    {"No.": "29.", "theme": "틀린 어법 찾기"},
    {"No.": "29.", "theme": "옳은 어법 선택"},
    {"No.": "30.", "theme": "어휘"},
    {"No.": "31.", "theme": "빈칸 추론(단어)"},
    {"No.": "32.", "theme": "빈칸 추론(문장)"},
    {"No.": "33.", "theme": "빈칸 추론(문장)"},
    {"No.": "34.", "theme": "빈칸 추론(문장)"},
    {"No.": "35.", "theme": "전체 흐름과 관계 없는 문장 파악"},
    {"No.": "36.", "theme": "글의 순서 추론"},
    {"No.": "37.", "theme": "글의 순서 추론"},
    {"No.": "38.", "theme": "문장 삽입"},
    {"No.": "39.", "theme": "문장 삽입"},
    {"No.": "40.", "theme": "문장 요약"},
    {"No.": "41.", "theme": "장문독해 - 제목"},
    {"No.": "42.", "theme": "장문독해 - 낱말의 쓰임"},
    {"No.": "43.", "theme": "장문독해 - 순서 배열"},
    {"No.": "44.", "theme": "장문독해 - 가리키는 대상"},
    {"No.": "45.", "theme": "장문 독해 - 내용 이해"}
    ]


    problem_No_theme_dict_list = [
    {"No.": "35.", "theme": "전체 흐름과 관계 없는 문장 파악"},
    {"No.": "36.", "theme": "글의 순서 추론"},
    {"No.": "38.", "theme": "문장 삽입"},
    {"No.": "40.", "theme": "문장 요약"},
    ]


    problem_No_theme_dict_list = [
    {"No.": "18.", "theme": "글의 목적 파악"},
    {"No.": "19.", "theme": "심경 분위기 파악"},
    {"No.": "20.", "theme": "필자의 주장 파악"},
    {"No.": "21.", "theme": "괄호 친 문장의 의미 파악"},
    {"No.": "22.", "theme": "글의 요지 파악"},
    {"No.": "23.", "theme": "글의 주제 파악"},
    {"No.": "24.", "theme": "글의 제목 추론"},
    {"No.": "26.", "theme": "대상 불일치 파악"},
    {"No.": "28.", "theme": "틀린 어법 찾기"},
    {"No.": "29.", "theme": "옳은 어법 선택"},
    {"No.": "30.", "theme": "어휘"},
    {"No.": "31.", "theme": "빈칸 추론(단어)"},
    {"No.": "32.", "theme": "빈칸 추론(문장)"},
    {"No.": "33.", "theme": "빈칸 추론(문장)"},
    {"No.": "34.", "theme": "빈칸 추론(문장)"},
    {"No.": "35.", "theme": "전체 흐름과 관계 없는 문장 파악"},
    {"No.": "36.", "theme": "글의 순서 추론"},
    {"No.": "37.", "theme": "글의 순서 추론"},
    {"No.": "38.", "theme": "문장 삽입"},
    {"No.": "39.", "theme": "문장 삽입"},
    {"No.": "40.", "theme": "문장 요약"},
    ]



    problem_info_list = []

    for problem_No_theme in problem_No_theme_dict_list:

        section_dict = {}
        essential_list = []

        start_time = time.time()  # 코드 실행 전 시간

        while( not(  ( len(set(essential_list) - set(section_dict.keys())) == 0 ) and len(section_dict) > 0 ) ):

            section_dict = {}

            problem, few_shot_message_list = make_single_problem(problem_No_theme['theme'], temperature= temperature, top_p= top_p)

            # Use a regex to match each section
            for match in re.finditer(r'(질문|주어진 글|주어진 문장|문제 본문|요약 문장|선택지|해설|주요 어휘)\s*:\s*(.*?)(?=(질문|주어진 글|주어진 문장|문제 본문|요약 문장|선택지|해설|주요 어휘)\s*:\s*|\Z)', problem.strip(), re.DOTALL):
                key = match.group(1).strip()
                value = match.group(2).strip()
                section_dict[key] = value

            essential_list = ['질문','문제 본문','해설']

            if(problem_No_theme['theme'] == "틀린 어법 찾기" or problem_No_theme['theme'] == "어휘"
                    or problem_No_theme['theme'] == "전체 흐름과 관계 없는 문장 파악"):
                pass

            elif(problem_No_theme['theme'] == "글의 순서 추론"):
                essential_list.append("주어진 글")
                essential_list.append("선택지")

            elif(problem_No_theme['theme'] == "문장 삽입"):
                essential_list.append("주어진 문장")

            elif(problem_No_theme['theme'] == "문장 요약"):
                essential_list.append("요약 문장")
                essential_list.append("선택지")

            else:
                essential_list.append("선택지")

            if ( len(section_dict) > 0 ) :
                if ( len(set(essential_list) - set(section_dict.keys())) == 0 ): #질문, 문제 본문, 선택지, 해설지 모두가 다 section_dict.keys()에 존재해야함.
                    problem_info_list.append({"No.": problem_No_theme['No.'], "theme" :  problem_No_theme['theme'], "problem" : section_dict, "few_shot_message" : few_shot_message_list})
                    end_time = time.time()  # 코드 실행 후 시간

                    elapsed_time = end_time - start_time
                    print(problem_No_theme['No.'] + " " + problem_No_theme['theme'] + " 출제 완료! / 소요 시간 : " + str(elapsed_time))

                else:
                    #재출제 해야함
                    print("")

            else:
                print("")

    return problem_info_list


def main():

    problem_info_list = make_testpaper(temperature=1.2, top_p=0.75)

    save_excel_data(problem_info_list, temperature=1.2, top_p=0.75)

    print()



    for temp in numpy.arange(0.6, 1.1, 0.2):
        for top_p in numpy.arange(0.80, 1.01, 0.05):

            problem_info_list = make_testpaper(temperature= temp,top_p= top_p)

            save_excel_data(problem_info_list, temperature= temp,top_p= top_p)

            print()



if __name__ == '__main__':
    main()
